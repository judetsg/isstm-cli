# Saisie des notes d'examen
import math
import os
import statistics
from decimal import Decimal
import pprint
from tqdm import tqdm
import openpyxl
import pandas as pd
import pdb

import questionary
from rethinkdb import RethinkDB
from data import db, redis_client
from operations import utility_functions
from utilities import chooser

r = RethinkDB()
# connect to the Rethinkdb database
conn = r.connect(db=db)


def saisir_note_examen():
    # while True:
    """
    0. Choisir l'année universitaire
    1. Choisir le niveau
    2. Choisir le semestre
    3. Choisir l'EC
    4. Choisir si on veut saisir les codes ou les notes
    5. Revenir au menu principal
    """
    # 0. Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1. Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 2. choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)
    # print(f" Semestre: {(redis_client.get('semestre_choisi')).decode()}")

    # 3. choisir l'ec
    semestre_choisi = redis_client.get('semestre_choisi').decode()
    ec_choisi = chooser.selector('Choisissez un EC', utility_functions.select_ec(conn, semestre_choisi))
    redis_client.set('ec_choisi', ec_choisi)
    # print(f"EC: {(redis_client.get('ec_choisi')).decode()}")

    # choisir la session
    session = chooser.selector('Session', [
        questionary.Choice(title='Session 1', value=1),
        questionary.Choice(title='Session 2', value=2)
    ])
    redis_client.set('session', session)

    # 4. choisir si on veut saisir les codes ou les notes
    activite = chooser.selector('Type de Saisie', [
        questionary.Choice(title='Anonymat', value='saisie_code'),
        questionary.Choice(title='Correction Anonymat', value='correction_anonymat'),
        questionary.Choice(title='Note', value='saisie_note')
    ])
    # redis_client.set('activite', activite)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    ec = redis_client.get('ec_choisi').decode()
    session = redis_client.get('session').decode()

    if activite == 'saisie_code':

        # Saisir les lettres du code si on veut saisir les Anonymats
        lettre_code = questionary.text(message='Entrez les lettres du code').ask()
        # redis_client.set('lettre_code', lettre_code)

        # afficher la liste des étudiants en utilisant le niveau choisi
        while True:
            try:
                # Si session 2, inserer directement dans la table moyenne_ec et
                # filtrer uniquement les etudiants qui ont une ec à repasser
                os.system('clear')
                # breakpoint()
                if session == "2":
                    # recuperer la liste des étudiants ayant une matière à repasser
                    etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                         utility_functions.select_etudiant_repechage(conn, ec))

                    # saisir le rang
                    rang = questionary.text('Rang:').ask()

                    # inserer dans la table moyenne_ec directement
                    r.table('moyenne_ec').insert({
                        'id': r.uuid(),
                        'id_annee': annee,
                        'id_ec': ec,
                        'id_etudiant': etudiant_id,
                        'id_semestre': semestre,
                        'id_session': session,
                        'anonymat': lettre_code + rang,
                    }).run(conn)
                else:
                    etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                         utility_functions.select_etudiant(conn, niveau_choisi))

                    # saisir le rang
                    rang = questionary.text('Rang:').ask()

                    # saisir les anonymats
                    """
                    Pour saisir les anonymats, il nous faut l'année universitaire, le niveau, l'ec, le semestre et l'id de l'etudiant.
                    Nous allons utiliser et afficher les etudiants, en choisir un entrer son anonymat et ainsi de suite,
                    jusqu'à ce que nous cassons la boucle
                    """
                    r.table('notes').insert({
                        'id': r.uuid(),
                        'id_annee': annee,
                        'niveau': niveau,
                        'semestre': semestre,
                        'etudiant_id': etudiant_id,
                        'id_ec': ec,
                        'session': session,
                        'anonymat': lettre_code + rang,
                    }).run(conn)
            except KeyboardInterrupt:
                break

    if activite == 'correction_anonymat':

        # Saisir les lettres du code si on veut saisir les Anonymats
        lettre_code = questionary.text(message='Entrez les lettres du code').ask()
        # redis_client.set('lettre_code', lettre_code)

        # afficher la liste des étudiants en utilisant le niveau choisi
        while True:
            try:
                # Si session 2, inserer directement dans la table moyenne_ec et
                # filtrer uniquement les etudiants qui ont une ec à repasser
                os.system('clear')
                # breakpoint()
                if session == "2":
                    # recuperer la liste des étudiants ayant une matière à repasser
                    etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                         utility_functions.select_etudiant_repechage(conn, ec))

                    # saisir le rang
                    rang = questionary.text('Rang:').ask()

                    # inserer dans la table moyenne_ec directement
                    r.table('moyenne_ec')\
                        .filter(r.and_(
                        r.row['id_etudiant'].eq(etudiant_id),
                        r.row['id_ec'].eq(ec),
                        r.row['id_session'].eq("2"),
                        r.row['anonymat'].match(f'^{lettre_code}')
                    )).update({'anonymat': lettre_code + rang}).run(conn)
                else:
                    etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                         utility_functions.select_etudiant(conn, niveau_choisi))

                    # saisir le rang
                    rang = questionary.text('Rang:').ask()

                    # saisir les anonymats
                    """
                    Pour saisir les anonymats, il nous faut l'année universitaire, le niveau, l'ec, le semestre et l'id de l'etudiant.
                    Nous allons utiliser et afficher les etudiants, en choisir un entrer son anonymat et ainsi de suite,
                    jusqu'à ce que nous cassons la boucle
                    """
                    r.table('notes') \
                        .filter((r.row['etudiant_id'] == etudiant_id) & \
                                (r.row['anonymat'].match(f'^{lettre_code}')) & \
                                (r.row['id_ec'] == ec)) \
                        .update({
                        'anonymat': lettre_code + rang
                    }).run(conn)
            except KeyboardInterrupt:
                break

    if activite == 'saisie_note':
        """
        1. Récuperer les anonymats correspondants et les afficher en liste comme la liste des etudiants
        2. Saisir les notes correspondantes
        """
        while True:
            try:
                # Si session 2, inserer directement dans la table moyenne_ec et
                # filtrer uniquement les etudiants qui ont une ec à repasser
                os.system('clear')
                # breakpoint()
                if session == "2":
                    # recuperer la liste des étudiants ayant une matière à repasser
                    anonymat_id = chooser.tuple_selector('Choisir anonymat: ',
                                                         utility_functions.select_note_repechage(conn, annee,
                                                                                                 semestre, ec, session))

                    print(f"Anonymat: {anonymat_id}")
                    # saisir la note
                    note = questionary.text('Note:').ask()
                    try:
                        decimal_note = Decimal(note)
                        r.table('moyenne_ec').get(anonymat_id).update({
                            'note': float(decimal_note)
                        }).run(conn)
                    except ValueError:
                        print(f"Erreur: tapez un nombre ({ValueError})")
                else:
                    anonymat_id = chooser.tuple_selector('Choisir anonymat: ',
                                                         utility_functions.select_note(conn, annee, niveau, semestre, ec,
                                                                                       session))
                    print(f"Anonymat: {anonymat_id}")
                    # saisir la note
                    note = questionary.text('Note:').ask()
                    try:
                        decimal_note = Decimal(note)
                        r.table('notes').get(anonymat_id).update({
                            'note': float(decimal_note),
                            'type': 'EX'
                        }).run(conn)
                    except ValueError:
                        print(f"Erreur: tapez un nombre ({ValueError})")
            except KeyboardInterrupt:
                break


def quitter():
    print("Tapez Ctrl+C pour quitter")


def saisir_note(type):
    """
        0. Choisir l'année universitaire
        1. Choisir le niveau
        2. Choisir le semestre
        3. Choisir l'EC
        4. Saisir les notes
        5. Revenir au menu principal
        """
    # 0. Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1. Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 2. choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)
    # print(f" Semestre: {(redis_client.get('semestre_choisi')).decode()}")

    # 3. choisir l'ec
    semestre_choisi = redis_client.get('semestre_choisi').decode()
    ec_choisi = chooser.selector('Choisissez un EC', utility_functions.select_ec(conn, semestre_choisi))
    redis_client.set('ec_choisi', ec_choisi)
    # print(f"EC: {(redis_client.get('ec_choisi')).decode()}")

    # choisir la session
    session = chooser.selector('Session', [
        questionary.Choice(title='Session 1', value=1),
        questionary.Choice(title='Session 2', value=2)
    ])
    # session = '1'  # il n'y a pas de note de CC à la session 2
    redis_client.set('session', session)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    ec = redis_client.get('ec_choisi').decode()
    session = redis_client.get('session').decode()

    # Saisir les notes
    # afficher la liste des étudiants en utilisant le niveau choisi
    while True:
        try:
            os.system('clear')
            if session == "2":
                # recuperer la liste des étudiants ayant une matière à repasser
                etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                     utility_functions.select_etudiant_repechage(conn, ec))

                # saisir la note
                note = questionary.text('Note:').ask()
                decimal_note = Decimal(note)

                # inserer dans la table moyenne_ec directement
                r.table('moyenne_ec').insert({
                    'id': r.uuid(),
                    'id_annee': annee,
                    'id_ec': ec,
                    'id_etudiant': etudiant_id,
                    'id_semestre': semestre,
                    'id_session': session,
                    'note': float(decimal_note),
                }).run(conn)

            else:
                etudiant_id = chooser.tuple_selector('Choisir un étudiant: ',
                                                     utility_functions.select_etudiant(conn, niveau_choisi))

                # saisir la note
                note = questionary.text('Note:').ask()
                decimal_note = Decimal(note)

                # saisir les anonymats
                """
                Pour saisir les anonymats, il nous faut l'année universitaire, le niveau, l'ec, le semestre et l'id de l'etudiant.
                Nous allons utiliser et afficher les etudiants, en choisir un entrer son anonymat et ainsi de suite,
                jusqu'à ce que nous cassons la boucle
                """

                r.table('notes').insert({
                    'id': r.uuid(),
                    'id_annee': annee,
                    'niveau': niveau,
                    'semestre': semestre,
                    'etudiant_id': etudiant_id,
                    'id_ec': ec,
                    'session': session,
                    'note': float(decimal_note),
                    'type': type
                }).run(conn)


        except KeyboardInterrupt:
            break


def calculer_moyenne_ec_session_1():
    """
    Pour calculer les matières à repasser à la premiere session d'un semestre, nous allons:
    1. prendre la liste des ECs qui ont des notes pour le semestre et l'année universitaire donnée
    2. pour chaque EC, pour chaque etudiant, calculer la moyenne
    3. pour chaque etudiant, assembler les moyennes des EC en un seul document et insérer dans la table moyenne_ec
    4. pour chaque étudiant, verifier si une EC est à repasser ou pas
    5. compiler la liste des matieres à repasser pour un semestre donné et un niveau donné en un fichier excel
    6. Sauvegarder le fichier excel
    :return:
    """

    # 1. prendre la liste des ECs qui ont des notes pour le semestre et l'année universitaire donnée
    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 1.2 choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    session = '1'

    # 1.3 obtenir la liste des ECs ayant des notes d'examen pour ce niveau et ce semestre
    ec_list = r.table('notes').group('id_ec').filter(
        r.and_(
            (r.row['id_annee'].eq(annee)),
            (r.row['niveau'].eq(niveau)),
            (r.row['semestre'].eq(semestre)),
            (r.row['session'].eq(session))
        )
    ).run(conn)
    # pprint.pprint(ec_list)
    # ec_list above is a dict like this: {'1':[{...}],'2':[{...}]}
    # 2. pour chaque EC, pour chaque etudiant, calculer la moyenne
    # effacer d'abord les notes precedentes s'il y a eu
    r.table('moyenne_ec').filter(
        r.and_(
            r.row['id_annee'].eq(annee),
            r.row['id_niveau'].eq(niveau),
            r.row['id_semestre'].eq(semestre),
            r.row['id_session'].eq('1')
        )
    ).delete().run(conn)

    for unique_ec, list_of_notes in tqdm(ec_list.items()):
        il_y_a_eu_des_CC = False
        # pour calculer la moyenne, verifions d'abord si il y a eu des CC pour cet EC
        for note in list_of_notes:
            for key, value in note.items():
                if value == "CC":
                    il_y_a_eu_des_CC = True
                    break

        # print(f'EC: {unique_ec} - Il y a eu des CC: {il_y_a_eu_des_CC}')
        # si il y a eu des CCs, compter combien de CCs il y a eu, ensuite, faire la moyenne de ces CCs
        # pour compter combien de CC il y a eu, nous allons compter le nombre d'occurence de (id_etudiant,type)
        # prenons chaque note et comptons l'occurence du tuple en dessous, ensuite nous allons prendre la valeur max

        list_of_CCs = []
        number_of_CCs = 0  # nombre de CCs pour l'EC
        if il_y_a_eu_des_CC:
            tuple_of_keys = ("etudiant_id", "type")
            # Iterate over the list of dictionaries
            for note in list_of_notes:
                # Initialize the count variable
                cc_count = 0
                # Check if the tuple of keys exists in the dictionary
                if all(key in note for key in tuple_of_keys):
                    # Check if the value exists in the dictionary
                    if note[tuple_of_keys[1]] == "CC":
                        # Increment the count variable
                        cc_count += 1

                list_of_CCs.append(cc_count)

        if len(list_of_CCs) != 0:
            number_of_CCs = max(list_of_CCs)

        # pour chaque étudiant, s'il y a eu des CCs, obtenir une liste des notes de CCs et faire la moyenne
        list_etudiants = utility_functions.select_etudiant(conn,
                                                           niveau_choisi)  # this returns a list of tuple of students
        # like this [(id, nom, prenoms)]
        for etudiant in list_etudiants:  # pour chaque etudiant
            if number_of_CCs != 0:  # s'il y a eu des CCs
                moyenne_CCs = 0  # la moyenne des CCs
                note_EX = 0  # la note aux examems
                liste_note_CCs = []  # liste des notes de CCs
                for note in list_of_notes:  # pour chaque note
                    if note['etudiant_id'] == etudiant[0] and note[
                        'type'] == "CC":  # comparer si c'est la note de l'etudiant concerné
                        liste_note_CCs.append(note['note'])  # ajouter la note de CCs à la liste

                    if note['etudiant_id'] == etudiant[0] and note['type'] == "EX":
                        note_EX = note['note']

                if len(liste_note_CCs) != 0:  # si l'etudiant a assisté à tous les CCs
                    moyenne_CCs = round(sum(liste_note_CCs) / number_of_CCs,
                                        2)  # arrondir la moyenne à 2 chiffres après la virgule

                # calculer la moyenne de l'ec
                moyenne_EC = note_EX * 0.7 + moyenne_CCs * 0.3
                ec_a_repasser_avec_cc = False
                if moyenne_EC < 10:
                    ec_a_repasser_avec_cc = True
                # inserer la moyenne ec pour l'etudiant
                r.table('moyenne_ec').insert({
                    'id_annee': annee,
                    'id_niveau': niveau,
                    'id_semestre': semestre,
                    'id_session': session,
                    'id_etudiant': etudiant[0],
                    'id_ec': unique_ec,
                    'note': moyenne_EC,
                    'repasser': ec_a_repasser_avec_cc
                }).run(conn)
            else:
                # inserer la moyenne EC directement
                note_EX = 0
                for note in list_of_notes:
                    if note['etudiant_id'] == etudiant[0]:
                        if note.get('note') is not None:
                            note_EX = note['note']
                ec_a_repasser = False
                if note_EX < 10:
                    ec_a_repasser = True
                r.table('moyenne_ec').insert({
                    'id_annee': annee,
                    'id_niveau': niveau,
                    'id_semestre': semestre,
                    'id_session': session,
                    'id_etudiant': etudiant[0],
                    'id_ec': unique_ec,
                    'note': note_EX,
                    'repasser': ec_a_repasser
                }).run(conn)


def generer_matiere_a_repasser_session_1():
    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 1.2 choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    session = '1'

    # 5. compiler la liste des matieres à repasser pour un semestre donné et un niveau donné en un fichier excel
    # 6. Sauvegarder le fichier excel
    # `liste_moyenne_ec` est un dict avec chaque id_etudiant comme key et une liste de dict de chaque note que cet
    # étudiant a obtenu pour la session 1 du semestre, niveau et année universitaire donnée, comme valeur.
    liste_moyenne_ec = r.table('moyenne_ec').group('id_etudiant').filter(
        r.and_(
            (r.row['id_annee'].eq(annee)),
            (r.row['id_niveau'].eq(niveau)),
            (r.row['id_semestre'].eq(semestre)),
            (r.row['id_session'].eq(session))
        )
    ).run(conn)

    # nous allons d'abord obtenir la liste des Ecs avec des notes
    liste_ec = r.table('moyenne_ec').group('id_ec').filter(
        r.and_(
            (r.row['id_annee'].eq(annee)),
            (r.row['id_niveau'].eq(niveau)),
            (r.row['id_semestre'].eq(semestre)),
            (r.row['id_session'].eq(session))
        )
    ).run(conn)

    ECs = []  # la liste des ECs ayant des notes pour l'année, le niveau, le semestre et la session donnée
    for ec, value in liste_ec.items():
        ECs.append(ec)

    # obtenir les "pretty names" (silly comments, I know, I hope nobody would have to read this comment one day
    niveau_pretty = utility_functions.get_pretty_name(conn, 'niveau', niveau, 'niveau')
    semestre_pretty = utility_functions.get_pretty_name(conn, 'semestre', semestre, 'semestre')

    # comme header nous avons | niveau | semestre | session | id_etudiant | id_ec | note |
    headers = ('niveau', 'semestre', 'session', 'matricule', 'nom', 'prenoms', 'ec', 'note')

    repertoire_actuel = os.getcwd()  # là ou on se trouve
    repertoire_xlsx = os.path.join(repertoire_actuel, 'xlsx')  # là ou se trouve le repertoire xlsx
    nom_du_ficher_excel = f'matiere_a_repasser_Semestre{semestre_pretty}_Niveau{niveau_pretty}_Annee{annee}.xlsx'
    nom_du_ficher_excel_avec_note = f'note_Semestre{semestre_pretty}_Niveau{niveau_pretty}_Annee{annee}.xlsx'

    # créer un fichier excel
    matiere_a_repasser = openpyxl.Workbook()
    note = openpyxl.Workbook()

    # créer un objet excel
    classeur = matiere_a_repasser.active
    classeur_note = note.active
    # inserer les headers
    classeur.append(headers)
    classeur_note.append(headers)

    # nous allons utiliser la liste_moyenne_ec et inserer les notes dans le fichier excel
    for etudiant_id, notes in tqdm(liste_moyenne_ec.items()):
        matricule = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'matricule')
        nom = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'nom')
        prenoms = None
        try: # des fois un étudiant n'a pas de prénoms
            prenoms = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'prenoms')
        except:
            pass

        for note_dict in notes:
            ec = utility_functions.get_pretty_name(conn, 'element_const', note_dict['id_ec'], 'appellation_ec')
            if note_dict['note'] >= 10:
                row_to_insert = (niveau_pretty, semestre_pretty, session , matricule, nom, prenoms, ec, "O")
            else:
                row_to_insert = (niveau_pretty, semestre_pretty, session, matricule, nom, prenoms, ec, "X")
            classeur.append(row_to_insert)
            note_to_insert = (niveau_pretty, semestre_pretty, session , matricule, nom, prenoms, ec, note_dict['note'])
            classeur_note.append(note_to_insert)
            # pdb.set_trace()

    matiere_a_repasser.save(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    note.save(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    # nous allons créer un 'pivot table' à partir du fichier excel
    # Créer un dataframe pandas
    data = pd.read_excel(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    data_note = pd.read_excel(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    # df = pd.DataFrame(sheet.values)
    data.fillna('', inplace=True)
    data_note.fillna(0, inplace=True)

    # pprint.pprint(data)

    # Créer un tableau croisé à partir du DataFrame
    pivot_table = pd.pivot_table(data, values='note', index= ['matricule', 'nom', 'prenoms'], columns='ec',
                                 aggfunc=lambda x: ''.join(str(v) for v in x),
                                 fill_value='')
    pivot_table_note = pd.pivot_table(data_note, values='note', index=['matricule', 'nom', 'prenoms'], columns='ec',
                                 aggfunc=sum,
                                 fill_value=0)
    # pivot_table['note'] = pivot_table['note'].astype(str)
    # pprint.pprint(pivot_table)

    # exporter le pivot table vers excel
    writer = pd.ExcelWriter(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    writer_note = pd.ExcelWriter(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    pivot_table.to_excel(writer, sheet_name='PivotTable')
    pivot_table_note.to_excel(writer_note, sheet_name='PivotTable')

    # Sauvegarder
    writer.close()
    writer_note.close()


def calculer_moyenne_ue_session_1():
    """
    Pour chaque ue, dans un semestre donné, nous allons faire la moyenne de chaque ec.
    La moyenne de chaque ue du semestre donnera la moyenne du semestre.
    Si un étudiant obtient la moyenne pour chaque ue, il obtient le semestre
    :return:
    """

    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 1.2 choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    session = '1'

    # Obtenir la liste des UE pour un semestre
    liste_ue_par_semestre = r.table('unite_ens').filter(
        r.row['semestre_id'].eq(semestre)
    ).run(conn)
    liste_ue = []
    for ue in liste_ue_par_semestre:
        liste_ue.append(ue['id'])

    # Effacer les moyennes UE precedentes si necessaire
    # effacer d'abord les notes precedentes s'il y a eu
    r.table('moyenne_ue').filter(
        r.and_(
            r.row['id_annee'].eq(annee),
            r.row['id_niveau'].eq(niveau),
            r.row['id_semestre'].eq(semestre),
            r.row['id_session'].eq('1')
        )
    ).delete().run(conn)

    # Obtenir la liste des étudiants pour un niveau
    list_etudiants = utility_functions.select_etudiant(conn,niveau)
    # pprint.pprint(list_etudiants)

    for etudiant in tqdm(list_etudiants):
        # pdb.set_trace()
        # Obtenir la liste des EC pour chaque UE et calculer la moyenne
        for ue_id in liste_ue:
            liste_note_ec = []
            liste_ec_par_ue = r.table('element_const').filter(
                r.row['ue_id'].eq(ue_id)
            ).run(conn)
            # print(f'Liste des EC de {ue} : {liste_ec_par_ue}')

            # pour chaque étudiant, obtenir la liste des moyennes des EC pour un UE
            # puis faire la moyenne pour avoir la moyenne de l'UE
            for ec in liste_ec_par_ue:
                # print(f'Recuperation de la note de {ec}')
                moyenne_ec = r.table('moyenne_ec').filter(
                    r.and_(
                        r.row['id_ec'].eq(ec['id']),
                        r.row['id_etudiant'].eq(etudiant[0]),
                        r.row['id_semestre'].eq(semestre)
                    )
                ).run(conn)

                # il doit y avoir normalement un seul objet:
                # pdb.set_trace()
                # note = 0
                for value in moyenne_ec:
                    note = value['note']
                    # break

                liste_note_ec.append(note)
                # pdb.set_trace()

            # pdb.set_trace()
            # calculer la moyenne de l'ue
            # pprint.pprint(liste_note_ec)
            if len(liste_note_ec) != 0:
                moyenne_ue = statistics.mean(liste_note_ec)
                ue_valide = False
                if moyenne_ue >= 10:
                    ue_valide = True
                # print(f"Etudiant:{etudiant[1]} - UE:{ue['id']} - Moy: {moyenne_ue}")

                # inserer dans la base
                r.table('moyenne_ue').insert({
                    'id_annee': annee,
                    'id_niveau': niveau,
                    'id_semestre': semestre,
                    'id_session': session,
                    'id_ue': ue_id,
                    'id_etudiant': etudiant[0],
                    'moyenne_ue': moyenne_ue,
                    'valide': ue_valide
                }).run(conn)

def admission_session_1():
    """
    Calcul l'admission à la premiere session. Si tous les ue d'un semestre sont validé, un étudiant passe à la première
    session.
    :return:
    """
    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()

    # effacer les données precedentes s'il y a eu
    r.table('admission_session1').filter(
        r.and_(
            r.row['id_annee'].eq(annee),
            r.row['id_niveau'].eq(niveau)
        )
    ).delete().run(conn)

    # Pour chaque etudiant, verifier si tous les semestres sont validés.
    # Obtenir la liste des étudiants pour un niveau
    list_etudiants = utility_functions.select_etudiant(conn, niveau)
    for etudiant in tqdm(list_etudiants):
        # Obtenir la liste des moyennes des UE
        liste_moyenne_ue = r.table('moyenne_ue').eq_join('id_ue', r.table('unite_ens')).without(
            {'right': 'id'}
        ).zip().eq_join('id_etudiant', r.table('etudiant')).without({'right': 'id'}).zip().filter(
            r.and_(
                r.row['id_etudiant'].eq(etudiant[0]),
                r.row['id_niveau'].eq(niveau)
            )).run(conn)

        moyennes_ue = []
        for moyenne_ue in liste_moyenne_ue:
            moyennes_ue.append(moyenne_ue)

        admis_session1 = True
        for moyenne_ue in moyennes_ue:
            # pdb.set_trace()
            if moyenne_ue['valide'] == False:
                admis_session1 = False
                break

        # pdb.set_trace()
        if admis_session1 == True:
            # calculer la moyenne génerale
            moyennes = []
            for moyenne_ue in moyennes_ue:
                moyennes.append(moyenne_ue['moyenne_ue'])

            # pdb.set_trace()
            moyenne_generale = statistics.mean(moyennes)

            # calculer la mention
            mention = utility_functions.get_mention(moyenne_generale)

            # inserer dans la base
            r.table('admission_session1').insert({
                'id_annee': annee,
                'id_niveau': niveau,
                'id_etudiant': etudiant[0],
                'admis_session1': admis_session1,
                'moyenne_generale': moyenne_generale,
                'mention': mention
            }).run(conn)


def calculer_moyenne_ue_session_2():
    """
    Pour chaque ue, dans un semestre donné, nous allons faire la moyenne de chaque ec.
    La moyenne de chaque ue du semestre donnera la moyenne du semestre.
    Si un étudiant obtient la moyenne pour chaque ue, il obtient le semestre
    :return:
    """

    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 1.2 choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    session = '2'

    # Obtenir la liste des UE pour un semestre
    liste_ue_par_semestre = r.table('unite_ens').filter(
        r.row['semestre_id'].eq(semestre)
    ).run(conn)
    liste_ue = []
    for ue in liste_ue_par_semestre:
        liste_ue.append(ue['id'])

    # Effacer les moyennes UE precedentes si necessaire
    # effacer d'abord les notes precedentes s'il y a eu
    r.table('moyenne_ue_session_2').filter(
        r.and_(
            r.row['id_annee'].eq(annee),
            r.row['id_niveau'].eq(niveau),
            r.row['id_semestre'].eq(semestre),
            r.row['id_session'].eq('2')
        )
    ).delete().run(conn)

    # Obtenir la liste des étudiants pour un niveau
    list_etudiants = utility_functions.select_etudiant(conn,niveau)
    # pprint.pprint(list_etudiants)

    for etudiant in tqdm(list_etudiants):
        # pdb.set_trace()
        # Obtenir la liste des EC pour chaque UE et calculer la moyenne
        for ue_id in liste_ue:
            # ignorer les UE suivantes (MWAHAHAHAHAHA - fou rire maléfique)
            ue_a_ignorer = []
            if niveau == '1':
                ue_a_ignorer = ['4', '6', '9', '11', '7', '34']
            elif niveau == '2':
                ue_a_ignorer = ['35', '20', '26', '22', '36', '17']

            if ue_id in ue_a_ignorer:
                pass
            else:
                liste_note_ec = []
                liste_ec_par_ue = r.table('element_const').filter(
                    r.row['ue_id'].eq(ue_id)
                ).run(conn)
                # print(f'Liste des EC de {ue} : {liste_ec_par_ue}')

                # pour chaque étudiant, obtenir la liste des moyennes des EC pour un UE
                # puis faire la moyenne pour avoir la moyenne de l'UE
                for ec in liste_ec_par_ue:
                    note = 0
                    # print(f'Recuperation de la note de {ec}')
                    # si l'ec était à refaire, prendre la note, sinon mettre zero
                    # vérifier si l'ec était à refaire
                    a_refaire = False
                    verifier_si_a_refaire = r.table('moyenne_ec').filter(
                        r.and_(
                            r.row['id_ec'].eq(ec['id']),
                            r.row['id_etudiant'].eq(etudiant[0]),
                            r.row['id_semestre'].eq(semestre),
                            r.row['id_session'].eq('1') # la verification des notes à refaire se fait dans la session 1 uniquement
                        )
                    ).run(conn)
                    # pprint.pp(verifier_si_a_refaire)
                    for value in verifier_si_a_refaire:
                        a_refaire = value['repasser']

                    # si a_refaire, prendre la note de la deuxieme session, sinon prendre la note de la session 1
                    if a_refaire:
                        moyenne_ec = r.table('moyenne_ec').filter(
                            r.and_(
                                r.row['id_ec'].eq(ec['id']),
                                r.row['id_etudiant'].eq(etudiant[0]),
                                r.row['id_semestre'].eq(semestre),
                                r.row['id_session'].eq('2')
                            )
                        ).run(conn)
                        # pprint.pp(moyenne_ec)
                        for value in moyenne_ec:
                            # breakpoint()
                            # check if the key 'note' exists (si le prof n'a pas encore remis de copie, par exemple)
                            # Franglish, quand tu me prends BWAHAHAHAHAHAH
                            if value.get('note') == None:
                                note = 0
                            else:
                                note = value['note']
                    else:
                        #prendre la note de la session 1
                        # breakpoint()
                        note = utility_functions.get_note_session1(conn, ec['id'],etudiant[0], semestre)

                    # breakpoint()
                    liste_note_ec.append(note)

                # calculer la moyenne de l'ue
                # pprint.pprint(liste_note_ec)
                # breakpoint()
                if len(liste_note_ec) != 0:
                    moyenne_ue = statistics.mean(liste_note_ec)
                    ue_valide = False
                    if moyenne_ue >= 10:
                        ue_valide = True
                    print(f"Etudiant:{etudiant[1]} - UE:{ue_id} - Moy: {moyenne_ue}")
                    # breakpoint()
                    # inserer dans la base
                    r.table('moyenne_ue_session_2').insert({
                        'id_annee': annee,
                        'id_niveau': niveau,
                        'id_semestre': semestre,
                        'id_session': session,
                        'id_ue': ue_id,
                        'id_etudiant': etudiant[0],
                        'moyenne_ue': moyenne_ue,
                        'valide': ue_valide
                    }).run(conn)


def admission_session_2():
    """
    Calcul l'admission à la premiere session. Si tous les ue d'un semestre sont validé, un étudiant passe à la première
    session.
    :return:
    """
    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauvegarder sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()

    # effacer les données precedentes s'il y a eu
    r.table('admission_session2').filter(
        r.and_(
            r.row['id_annee'].eq(annee),
            r.row['id_niveau'].eq(niveau)
        )
    ).delete().run(conn)

    # Pour chaque etudiant, verifier si tous les semestres sont validés.
    # Obtenir la liste des étudiants pour un niveau
    list_etudiants = utility_functions.select_etudiant(conn, niveau)
    for etudiant in tqdm(list_etudiants):
        # Obtenir la liste des moyennes des UE
        liste_moyenne_ue = r.table('moyenne_ue_session_2').eq_join('id_ue', r.table('unite_ens')).without(
            {'right': 'id'}
        ).zip().eq_join('id_etudiant', r.table('etudiant')).without({'right': 'id'}).zip().filter(
            r.and_(
                r.row['id_etudiant'].eq(etudiant[0]),
                r.row['id_niveau'].eq(niveau)
            )).run(conn)

        moyennes_ue = []
        for moyenne_ue in liste_moyenne_ue:
            moyennes_ue.append(moyenne_ue)

        # compter d'abord le nombre de moyenne d'UE
        # ensuite vérifier si le nombre d'UE validé est supérieur à 75%
        # si c'est le cas, l'etudiant est admis avec dette, sinon l'étudiant est recalé

        # compter le nombre de moyenne UE et definir le seuil d'admission avec dette
        nombre_moyenne_ue = len(moyennes_ue)
        # breakpoint()
        seuil_admission_avec_dette = int(math.floor(nombre_moyenne_ue * 0.7))

        # compter le nombre d'UE valide
        nombre_ue_valide = 0
        admis_session2 = False
        type_admission = None
        for moyenne_ue in moyennes_ue:
            # pdb.set_trace()
            if moyenne_ue['valide'] == True:
                nombre_ue_valide = nombre_ue_valide + 1

        # verifier admission
        # breakpoint()
        if nombre_ue_valide == nombre_moyenne_ue:
            # admis
            admis_session2 = True
            type_admission = "Admis"
        elif nombre_ue_valide >= seuil_admission_avec_dette:
            # admis avec dette
            admis_session2 = True
            type_admission = "Admis avec dette"
        else:
            # recalé
            admis_session2 = False
            type_admission = "Recalé"

        # print(f'Etudiant: {etudiant[0]} - Nbre UE Valide: {nombre_ue_valide}')
        # Calculer moyenne générale
        # pdb.set_trace()
        moyennes = []
        for moyenne_ue in moyennes_ue:
            moyennes.append(moyenne_ue['moyenne_ue'])

        # pdb.set_trace()
        moyenne_generale = statistics.mean(moyennes)

        # calculer la mention
        mention = utility_functions.get_mention(moyenne_generale)

        # inserer dans la base
        r.table('admission_session2').insert({
            'id_annee': annee,
            'id_niveau': niveau,
            'id_etudiant': etudiant[0],
            'admis_session2': admis_session2,
            'moyenne_generale': moyenne_generale,
            'type_admission': type_admission,
            'mention': mention,
            'nombre_ue_valide': f'{nombre_ue_valide}/{nombre_moyenne_ue}',
            'seuil_admission_ue': seuil_admission_avec_dette
        }).run(conn)


def generer_note_session_2():
    # 1.0 Choisir l'année universitaire
    annee_choices = utility_functions.select_annee(conn)
    annee_choisi = chooser.selector("Choisissez une année universitaire", annee_choices)
    redis_client.set('annee', annee_choisi)

    # 1.1 Choisir le niveau
    niveaux_choices = utility_functions.select_niveau(conn)
    niveau_choisi = chooser.selector("Choisissez un niveau", niveaux_choices)
    # Sauver sur redis
    redis_client.set('niveau_choisi', niveau_choisi)

    # 1.2 choisir le semestre
    niveau_choisi = redis_client.get('niveau_choisi').decode()
    semestre_choisi = chooser.selector('Choisissez un semestre', utility_functions.select_semestre(conn, niveau_choisi))
    redis_client.set('semestre_choisi', semestre_choisi)

    # recuperer les variables nécessaires sur redis
    annee = redis_client.get('annee').decode()
    niveau = redis_client.get('niveau_choisi').decode()
    semestre = redis_client.get('semestre_choisi').decode()
    session = '2'

    # 5. compiler la liste des matieres à repasser pour un semestre donné et un niveau donné en un fichier excel
    # 6. Sauvegarder le fichier excel

    # `liste_moyenne_ec` est un dict avec chaque id_etudiant comme key et une liste de dict de chaque note que cet
    # étudiant a obtenu pour la session 1 du semestre, niveau et année universitaire donnée, comme valeur.
    liste_moyenne_ec = r.table('moyenne_ec').group('id_etudiant').filter(
        r.and_(
            (r.row['id_annee'].eq(annee)),
            # (r.row['id_niveau'].eq(niveau)),
            (r.row['id_semestre'].eq(semestre)),
            (r.row['id_session'].eq('1'))
        )
    ).run(conn)

    # nous allons d'abord obtenir la liste des Ecs avec des notes
    liste_ec = r.table('moyenne_ec').group('id_ec').filter(
        r.and_(
            (r.row['id_annee'].eq(annee)),
            # (r.row['id_niveau'].eq(niveau)),
            (r.row['id_semestre'].eq(semestre)),
            (r.row['id_session'].eq('1'))
        )
    ).run(conn)
    # breakpoint()
    ECs = []  # la liste des ECs ayant des notes pour l'année, le niveau, le semestre et la session donnée
    for ec, value in liste_ec.items():
        ECs.append(ec)

    # obtenir les "pretty names" (silly comments, I know, I hope nobody would have to read this comment one day
    niveau_pretty = utility_functions.get_pretty_name(conn, 'niveau', niveau, 'niveau')
    semestre_pretty = utility_functions.get_pretty_name(conn, 'semestre', semestre, 'semestre')

    # comme header nous avons | niveau | semestre | session | id_etudiant | id_ec | note |
    headers = ('niveau', 'semestre', 'session', 'matricule', 'nom', 'prenoms', 'ec', 'note')

    repertoire_actuel = os.getcwd()  # là ou on se trouve
    repertoire_xlsx = os.path.join(repertoire_actuel, 'xlsx')  # là ou se trouve le repertoire xlsx
    # nom_du_ficher_excel = f'matiere_a_repasser_Semestre{semestre_pretty}_Niveau{niveau_pretty}_Annee{annee}_Session{session}.xlsx'
    nom_du_ficher_excel_avec_note = f'note_Semestre{semestre_pretty}_Niveau{niveau_pretty}_Annee{annee}_Session{session}.xlsx'

    # créer un fichier excel
    matiere_a_repasser = openpyxl.Workbook()
    note = openpyxl.Workbook()

    # créer un objet excel
    classeur = matiere_a_repasser.active
    classeur_note = note.active
    # inserer les headers
    classeur.append(headers)
    classeur_note.append(headers)

    # nous allons utiliser la liste_moyenne_ec et inserer les notes dans le fichier excel
    for etudiant_id, notes in tqdm(liste_moyenne_ec.items()):
        matricule = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'matricule')
        nom = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'nom')
        prenoms = None
        try: # des fois un étudiant n'a pas de prénoms
            prenoms = utility_functions.get_pretty_name(conn, 'etudiant', etudiant_id, 'prenoms')
        except:
            pass
        # breakpoint()
        for note_dict in notes:
            ec = utility_functions.get_pretty_name(conn, 'element_const', note_dict['id_ec'], 'appellation_ec')
            # si la note est validé en session 1
            valide_session1 = utility_functions.est_valide_session1(conn, note_dict['id_ec'], etudiant_id,
                                                                    niveau, semestre)
            # print(f'Etudiant: {nom} - EC: {ec} - Session 1 Valide: {valide_session1}')

            if not valide_session1:
                # breakpoint()
                note_session_2 = utility_functions.get_note_session2(conn, note_dict['id_ec'],
                                                                     etudiant_id, semestre)
                note_to_insert = (niveau_pretty, semestre_pretty, session, matricule, nom, prenoms, ec, note_session_2)
                classeur_note.append(note_to_insert)
            else:
                # breakpoint()
                if note_dict.get('note') == None: # note pas encore saisie
                    note_to_insert = (niveau_pretty, semestre_pretty, session,
                                      matricule, nom, prenoms, ec, 0)
                    classeur_note.append(note_to_insert)
                else:
                    note_to_insert = (niveau_pretty, semestre_pretty, session,
                                      matricule, nom, prenoms, ec, note_dict['note'])
                    classeur_note.append(note_to_insert)

    # matiere_a_repasser.save(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    note.save(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    # nous allons créer un 'pivot table' à partir du fichier excel
    # Créer un dataframe pandas
    # data = pd.read_excel(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    data_note = pd.read_excel(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    # df = pd.DataFrame(sheet.values)
    # data.fillna('', inplace=True)
    data_note.fillna(0, inplace=True)

    # pprint.pprint(data)

    # Créer un tableau croisé à partir du DataFrame
    # pivot_table = pd.pivot_table(data, values='note', index= ['matricule', 'nom', 'prenoms'], columns='ec',
    #                              aggfunc=lambda x: ''.join(str(v) for v in x),
    #                              fill_value='')
    pivot_table_note = pd.pivot_table(data_note, values='note', index=['matricule', 'nom', 'prenoms'], columns='ec',
                                 aggfunc=sum,
                                 fill_value=0)
    # pivot_table['note'] = pivot_table['note'].astype(str)
    # pprint.pprint(pivot_table)

    # exporter le pivot table vers excel
    # writer = pd.ExcelWriter(os.path.join(repertoire_xlsx, nom_du_ficher_excel))
    writer_note = pd.ExcelWriter(os.path.join(repertoire_xlsx, nom_du_ficher_excel_avec_note))

    # pivot_table.to_excel(writer, sheet_name='PivotTable')
    pivot_table_note.to_excel(writer_note, sheet_name='PivotTable')

    # Sauvegarder
    # writer.close()
    writer_note.close()
